import unittest
import os
import sys
import tempfile
from openpyxl import load_workbook

# 添加项目根目录到路径，以便导入app模块
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

from app.model import State, StateManager
from app.utils.export_to_excel import export_statechart_to_excel


class TestExportToExcel(unittest.TestCase):
    def setUp(self):
        """设置测试环境，创建一个简单的状态机"""
        # 创建根状态
        self.root_state = State(name="TrafficLight", transition="", lifecycle="", parent=None, children=[])
        self.state_manager = StateManager(self.root_state)
        
        # 添加InService状态
        self.in_service = State(
            name="InService",
            transition="[*] -> Red :: Start effect {\n    b = 0x1;\n}",
            lifecycle="enter {\n    a = 0;\n    b = 0;\n}\nenter abstract InServiceAbstractEnter /*\n    Abstract Operation\n*/",
            parent="TrafficLight",
            children=[]
        )
        self.state_manager.add_state(self.root_state, self.in_service)
        
        # 添加Red状态
        self.red_state = State(
            name="Red",
            transition="",
            lifecycle="during {\n    a = 0x1 << 2;\n}",
            parent="InService",
            children=[]
        )
        self.state_manager.add_state(self.in_service, self.red_state)
        
        # 添加Yellow状态
        self.yellow_state = State(
            name="Yellow",
            transition="",
            lifecycle="",
            parent="InService",
            children=[]
        )
        self.state_manager.add_state(self.in_service, self.yellow_state)
        
        # 添加Idle状态
        self.idle_state = State(
            name="Idle",
            transition="",
            lifecycle="",
            parent="TrafficLight",
            children=[]
        )
        self.state_manager.add_state(self.root_state, self.idle_state)
        
        # 设置根状态转移
        self.root_state.transition = "[*] -> InService;\nInService -> Idle :: Maintain;\n! * -> Idle : if [a >= 20];"
        
        # 设置变量定义
        self.state_manager.variable_definitions = "def int a = 0;\ndef int b = 0x0;"

    def test_export_to_excel(self):
        """测试导出到Excel功能"""
        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # 导出状态机到Excel
            export_statechart_to_excel(self.state_manager, tmp_path)
            # 检查文件是否存在
            self.assertTrue(os.path.exists(tmp_path), "Excel文件未被创建")
            
            # 加载Excel文件进行验证
            workbook = load_workbook(tmp_path)
            
            # 检查工作表是否存在
            self.assertIn("States", workbook.sheetnames, "应该存在States工作表")
            self.assertIn("Variables", workbook.sheetnames, "应该存在Variables工作表")
            self.assertIn("Transitions", workbook.sheetnames, "应该存在Transitions工作表")
            self.assertIn("Forced Transitions", workbook.sheetnames, "应该存在Forced Transitions工作表")
            
            # 检查States工作表内容
            states_sheet = workbook["States"]
            # 检查表头
            self.assertEqual(states_sheet.cell(row=1, column=1).value, "状态名称")
            
            # 检查TrafficLight状态
            traffic_light_found = False
            in_service_found = False
            red_found = False
            for row in range(2, states_sheet.max_row + 1):
                if states_sheet.cell(row=row, column=1).value == "TrafficLight":
                    traffic_light_found = True
                elif states_sheet.cell(row=row, column=1).value == "InService":
                    in_service_found = True
                    # 检查生命周期
                    enter_actions = states_sheet.cell(row=row, column=4).value
                    self.assertIn("a = 0", enter_actions)
                    self.assertIn("abstract InServiceAbstractEnter", enter_actions)
                elif states_sheet.cell(row=row, column=1).value == "Red":
                    red_found = True
                    # 检查生命周期
                    during_actions = states_sheet.cell(row=row, column=5).value
                    self.assertIn("a = 1 << 2", during_actions)
            
            self.assertTrue(traffic_light_found, "TrafficLight状态应该存在")
            self.assertTrue(in_service_found, "InService状态应该存在")
            self.assertTrue(red_found, "Red状态应该存在")
            
            # 检查Variables工作表内容
            variables_sheet = workbook["Variables"]
            # 检查表头
            self.assertEqual(variables_sheet.cell(row=1, column=1).value, "变量名")
            
            # 检查变量定义
            var_a_found = False
            var_b_found = False
            for row in range(2, variables_sheet.max_row + 1):
                if variables_sheet.cell(row=row, column=1).value == "a":
                    var_a_found = True
                    self.assertEqual(variables_sheet.cell(row=row, column=2).value, "int")
                elif variables_sheet.cell(row=row, column=1).value == "b":
                    var_b_found = True
                    self.assertEqual(variables_sheet.cell(row=row, column=2).value, "int")
            
            self.assertTrue(var_a_found, "变量a应该存在")
            self.assertTrue(var_b_found, "变量b应该存在")
            
            # 检查Transitions工作表内容
            transitions_sheet = workbook["Transitions"]
            # 检查表头
            self.assertEqual(transitions_sheet.cell(row=1, column=1).value, "所属状态")
            
            # 检查转移定义
            start_to_red_found = False
            for row in range(2, transitions_sheet.max_row + 1):
                state_name = transitions_sheet.cell(row=row, column=1).value
                src_state = transitions_sheet.cell(row=row, column=2).value
                dst_state = transitions_sheet.cell(row=row, column=3).value
                
                if state_name == "InService" and src_state == "INIT_STATE" and dst_state == "Red":
                    start_to_red_found = True
                    # 检查动作
                    action = transitions_sheet.cell(row=row, column=6).value
                    self.assertIn("b = 1", action)
            
            self.assertTrue(start_to_red_found, "从初始到Red的转移应该存在")
            
            # 检查Forced Transitions工作表内容
            forced_sheet = workbook["Forced Transitions"]
            # 检查表头
            self.assertEqual(forced_sheet.cell(row=1, column=1).value, "所属状态")
            
            # 检查强制转移
            forced_to_idle_found = False
            for row in range(2, forced_sheet.max_row + 1):
                state_name = forced_sheet.cell(row=row, column=1).value
                src_state = forced_sheet.cell(row=row, column=2).value
                dst_state = forced_sheet.cell(row=row, column=3).value
                condition = forced_sheet.cell(row=row, column=4).value
                
                if state_name == "TrafficLight" and src_state == "*" and dst_state == "Idle":
                    forced_to_idle_found = True
                    self.assertEqual(condition, "a >= 20")
            
            self.assertTrue(forced_to_idle_found, "强制转移到Idle应该存在")
            
        finally:
            # 清理临时文件
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

if __name__ == "__main__":
    unittest.main()
