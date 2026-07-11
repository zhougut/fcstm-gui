import tempfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from ..model import StateManager
from .ui_to_dsl import state_manager_to_dsl
from pyfcstm.dsl import parse_with_grammar_entry
from pyfcstm.model import parse_dsl_node_to_state_machine, State


def export_statechart_to_excel(state_manager: StateManager, file_path: str):
    """
    将状态机信息导出为Excel文档
    
    Args:
        state_manager: 状态管理器对象
        file_path: 导出文件路径
    """
    # 将StateManager转换为DSL
    dsl_content = state_manager_to_dsl(state_manager)
    # 解析DSL为StateMachine
    ast_node = parse_with_grammar_entry(dsl_content, entry_name='state_machine_dsl')
    state_machine = parse_dsl_node_to_state_machine(ast_node)

    # 创建工作簿
    wb = Workbook()

    # 创建状态工作表
    states_sheet = wb.active
    states_sheet.title = "States"

    # 设置状态表头
    headers = ["状态名称", "父状态", "状态类型", "进入动作(Enter)", "执行中动作(During)", "退出动作(Exit)"]
    for col, header in enumerate(headers, 1):
        cell = states_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # 添加状态信息
    row = 2
    for fcstm_state in state_machine.walk_states():
        # 基本状态信息
        states_sheet.cell(row=row, column=1).value = fcstm_state.name
        states_sheet.cell(row=row, column=2).value = fcstm_state.parent.name if fcstm_state.parent else ""

        # 状态类型
        if len(fcstm_state.substate_name_to_id) > 0:
            states_sheet.cell(row=row, column=3).value = "复合状态"
        else:
            states_sheet.cell(row=row, column=4).value = "简单状态"

        # 生命周期动作
        enter_actions = []
        for enter in fcstm_state.on_enters:
            if hasattr(enter, 'is_abstract') and enter.is_abstract:
                enter_actions.append(f"abstract {enter.name}")
            else:
                ops = [f"{op.var_name} = {op.expr}" for op in enter.operations] if hasattr(enter, 'operations') else []
                enter_actions.append("\n".join(ops))
        states_sheet.cell(row=row, column=4).value = "\n".join(enter_actions)

        during_actions = []
        for during in fcstm_state.on_durings:
            if hasattr(during, 'is_abstract') and during.is_abstract:
                during_actions.append(f"abstract {during.name}")
            else:
                ops = [f"{op.var_name} = {op.expr}" for op in during.operations] if hasattr(during, 'operations') else []
                during_actions.append("\n".join(ops))
        states_sheet.cell(row=row, column=5).value = "\n".join(during_actions)

        exit_actions = []
        for exit in fcstm_state.on_exits:
            if hasattr(exit, 'is_abstract') and exit.is_abstract:
                exit_actions.append(f"abstract {exit.name}")
            else:
                ops = [f"{op.var_name} = {op.expr}" for op in exit.operations] if hasattr(exit, 'operations') else []
                exit_actions.append("\n".join(ops))
        states_sheet.cell(row=row, column=6).value = "\n".join(exit_actions)

        row += 1

    # 调整列宽
    for col in range(1, len(headers) + 1):
        states_sheet.column_dimensions[chr(64 + col)].width = 30

    # 创建变量工作表
    variables_sheet = wb.create_sheet("Variables")

    # 设置变量表头
    var_headers = ["变量名", "类型", "初始值"]
    for col, header in enumerate(var_headers, 1):
        cell = variables_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # 添加变量信息
    row = 2
    for var_def in state_machine.defines.values():
        variables_sheet.cell(row=row, column=1).value = var_def.name
        variables_sheet.cell(row=row, column=2).value = var_def.type
        variables_sheet.cell(row=row, column=3).value = str(var_def.init)
        row += 1

    # 调整列宽
    for col in range(1, len(var_headers) + 1):
        variables_sheet.column_dimensions[chr(64 + col)].width = 20

    # 创建转移工作表
    transitions_sheet = wb.create_sheet("Transitions")

    # 设置转移表头
    trans_headers = ["所属状态", "源状态", "目标状态", "事件", "条件", "动作"]
    for col, header in enumerate(trans_headers, 1):
        cell = transitions_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # 添加转移信息
    row = 2
    for fcstm_state in state_machine.walk_states():
        if fcstm_state.transitions:
            for transition in fcstm_state.transitions:
                # 所属状态
                transitions_sheet.cell(row=row, column=1).value = fcstm_state.name

                # 源状态
                if transition.from_state == "[*]":
                    transitions_sheet.cell(row=row, column=2).value = "[初始]"
                else:
                    transitions_sheet.cell(row=row, column=2).value = str(transition.from_state)

                # 目标状态
                if transition.to_state == "[*]":
                    transitions_sheet.cell(row=row, column=3).value = "[终止]"
                else:
                    transitions_sheet.cell(row=row, column=3).value = str(transition.to_state)

                # 事件
                if transition.event:
                    transitions_sheet.cell(row=row, column=4).value = transition.event.name

                # 条件
                if transition.guard:
                    transitions_sheet.cell(row=row, column=5).value = str(transition.guard)

                # 动作（效果）
                if transition.effects:
                    effects = [str(op.to_ast_node()) for op in transition.effects]
                    transitions_sheet.cell(row=row, column=6).value = "\n".join(effects)

                row += 1

    # 调整列宽
    for col in range(1, len(trans_headers) + 1):
        transitions_sheet.column_dimensions[chr(64 + col)].width = 25

    # 创建强制转移工作表
    forced_sheet = wb.create_sheet("Forced Transitions")

    # 设置强制转移表头
    forced_headers = ["所属状态", "源状态", "目标状态", "条件", "动作"]
    for col, header in enumerate(forced_headers, 1):
        cell = forced_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # 识别并添加强制转移信息
    row = 2
    for state in state_manager.get_all_states():
        for transition in state.transitions:
            source = transition.get("source", "")
            if not source.startswith("!"):
                continue

            forced_sheet.cell(row=row, column=1).value = state.get_full_path()
            forced_sheet.cell(row=row, column=2).value = source[1:].strip()
            forced_sheet.cell(row=row, column=3).value = transition.get("target", "")
            forced_sheet.cell(row=row, column=4).value = transition.get("condition", "")
            forced_sheet.cell(row=row, column=5).value = transition.get("action", "")
            row += 1

    # 调整列宽
    for col in range(1, len(forced_headers) + 1):
        forced_sheet.column_dimensions[chr(64 + col)].width = 25

    # 保存工作簿
    wb.save(file_path)
